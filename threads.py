import pandas as pd
import threading
import matplotlib.pyplot as plt
import ee
import numpy as np
from utils import calculate_relative_humidity, calculate_specific_humidity
# from model import preprocess_data
from openpyxl.styles import PatternFill, Font
from openpyxl.utils.dataframe import dataframe_to_rows
import matplotlib.pyplot as plt
from io import BytesIO
from openpyxl.drawing.image import Image
from openpyxl.worksheet.worksheet import Worksheet
import pandas as pd
import openpyxl.styles


class DataFetcher(threading.Thread):
    def __init__(self, queue, start_date, end_date, latitude, longitude, threadlock):
        super().__init__()
        self.queue = queue
        self.start_date = start_date
        self.end_date = end_date
        self.latitude = latitude
        self.longitude = longitude
        self.threadlock = threadlock


    def run(self):
        print("Initializing data fetch process...")

        self.threadlock.acquire()

        climateVariables = [
    'temperature_2m',
    'temperature_2m_min',
    'temperature_2m_max',
    'dewpoint_temperature_2m',
    'surface_pressure',
    'u_component_of_wind_10m',
    'v_component_of_wind_10m',
    'total_precipitation_sum',
    'soil_temperature_level_1',
    'total_evaporation_sum',
    'evaporation_from_bare_soil_sum',
    'evaporation_from_vegetation_transpiration_sum'
]
        point = ee.Geometry.Point([self.longitude, self.latitude])

        print("Fetching data for ", point.getInfo(), '...')
        merged_data = {}

        for variable in climateVariables:
            print(f"Fetching data for variable: {variable}...")
            dataset = (
                ee.ImageCollection('ECMWF/ERA5_LAND/DAILY_AGGR')
                .filterDate(self.start_date, self.end_date)
                .select(variable)
                .getRegion(point, 1000)
            )
            data = dataset.getInfo()
            print(f"Data for {variable} fetched successfully.")

            df = pd.DataFrame(data[1:], columns=data[0])
            if "time" not in merged_data:
                merged_data["time"] = df["time"]
            merged_data[variable] = df[variable]

        print("All variables fetched and merged.")
        merged_df = pd.DataFrame(merged_data)
        
        self.queue.put(merged_df)
        print("All data chunks fetched and added to queue.")
        self.threadlock.release()


class DataOrganizer(threading.Thread):
    def __init__(self, input_queue, output_queue, detailed_queue, predicted_queue, threadlock):
        super().__init__()
        self.input_queue = input_queue
        self.output_queue = output_queue
        self.detailed_queue = detailed_queue
        self.predicted_queue = predicted_queue
        self.threadlock = threadlock


    def run(self):
        print("Organizing data chunks into a single DataFrame...")
        self.threadlock.acquire()

        
        df = self.input_queue.get()
        weather_df = pd.DataFrame()

        weather_df['date'] = pd.to_datetime(df['time'].apply(lambda x: x / 1000), unit="s")
        weather_df['temperature'] = df['temperature_2m'] - 273.15
        weather_df['wind_speed'] = np.sqrt(df['u_component_of_wind_10m']**2 + df['v_component_of_wind_10m']**2)
        weather_df['wind_direction'] = np.rad2deg(np.arctan2(df['u_component_of_wind_10m'], df['v_component_of_wind_10m']))
        weather_df['total_precipitation'] = df['total_precipitation_sum']
        weather_df['relative_humidity'] = calculate_relative_humidity(df['temperature_2m'], df['dewpoint_temperature_2m'])
        weather_df['specific_humidity'] = calculate_specific_humidity(df['temperature_2m'], df['dewpoint_temperature_2m'], df['surface_pressure'])
        weather_df['soil_temperature'] = df['soil_temperature_level_1'] - 273.15

        print("Data Processsed and organized into a DataFrame!")


        self.detailed_queue.put((weather_df))
        csv = "predicted_weather.csv"
        ai_df = pd.read_csv(csv)
        ai_df['date'] = weather_df['date']
        self.predicted_queue.put(ai_df)

        print("Data organized into DataFrames and added to output queue!")

        self.threadlock.release()



class BaseWeatherWriter(threading.Thread):
    def __init__(self, output_queue, threadlock):
        super().__init__()
        self.output_queue = output_queue
        self.threadlock = threadlock

        
        
        self.seasons = {
            'Winter': [12, 1, 2],
            'Spring': [3, 4, 5],
            'Summer': [6, 7, 8],
            'Autumn': [9, 10, 11]
        }
        self.block_color = PatternFill(start_color="FFDDDDDD", end_color="FFDDDDDD", fill_type="solid")
    
    def prepare_data(self, weather_df):
        """Prepare data by adding month column"""
        weather_df['month'] = weather_df['date'].dt.month
        return weather_df
    
    def write_detailed_sheet(self, writer, seasonal_df, season):
        """Write detailed data sheet for a season"""
        detailed_sheet_name = f"{season}_Detailed"
        seasonal_df_copy = seasonal_df.copy()
        seasonal_df_copy.drop('month', axis=1, inplace=True)
        seasonal_df_copy.to_excel(writer, sheet_name=detailed_sheet_name, index=False)
    
    def write_aggregated_sheet(self, writer, seasonal_df, season):
        """Write aggregated statistics sheet for a season"""
        aggregated_sheet_name = f"{season}_Aggregated"
        workbook = writer.book
        worksheet = workbook.create_sheet(title=aggregated_sheet_name)
        
        stats_by_year = self.calculate_stats_by_year_and_variable(seasonal_df)
        start_row = 1
        
        for year, year_data in stats_by_year.groupby('year'):
            # Write headers
            worksheet.cell(row=start_row, column=1, value=f'Year:{year}')
            headers = ['Aggregation', 'Average', 'Standard Deviation', 'Minimum', 'Maximum']
            for col, header in enumerate(headers, start=2):
                worksheet.cell(row=start_row, column=col, value=header)
            worksheet.cell(row=start_row, column=1).font = Font(bold=True)
            
            start_row += 1
            
            # Write data
            for _, row in year_data.iterrows():
                for col_idx, value in enumerate(row, start=1):
                    cell = worksheet.cell(row=start_row, column=col_idx, value=value)
                    cell.fill = self.block_color
                start_row += 1
            start_row += 1
    
    def adjust_column_widths(self, workbook):
        """Adjust column widths for better readability"""
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                adjusted_width = max_length + 2
                worksheet.column_dimensions[column_letter].width = adjusted_width
    
    @staticmethod
    def calculate_stats_by_year_and_variable(df):
        """Calculate statistics by year for each variable"""
        df['month'] = df['date'].dt.month
        df['year'] = df['date'].dt.year

        weather_variables = {
            'temperature': 'Air Temperature (°C)',
            'relative_humidity': 'Humidity (°C)',
            'wind_speed': 'Wind Speed (m/s)',
            'wind_direction': 'Wind Direction (°)',
            'total_precipitation': 'Precipitation (mm)',
            'soil_temperature': 'Soil Temperature (°C)',
            # 'evaporation': 'Evaporation (mm)'
        }

        stats = []
        for column, variable_name in weather_variables.items():
            grouped_stats = df.groupby(['year']).agg(
                Average=(column, 'mean'),
                Standard_Deviation=(column, 'std'),
                Minimum=(column, 'min'),
                Maximum=(column, 'max')
            ).reset_index()
            grouped_stats['Weather Variable'] = variable_name
            stats.append(grouped_stats)

        stats_df = pd.concat(stats, ignore_index=True)
        stats_df = stats_df[['year', 'Weather Variable', 'Average', 'Standard_Deviation', 'Minimum', 'Maximum']]

        return stats_df


class RealWeatherWriter(BaseWeatherWriter):
    def run(self):
        self.threadlock.acquire()
        try:
            weather_df = self.output_queue.get()
            self.output_queue.put(weather_df.copy())
            print("Writing real weather data into Excel file with seasonal optimizations...")
            
            weather_df = self.prepare_data(weather_df)
            
            with pd.ExcelWriter('real_weather_data_seasonal.xlsx', engine='openpyxl') as writer:
                weather_df.to_excel(writer, sheet_name='Real Weather Data', index=False)
                
                for season, months in self.seasons.items():
                    seasonal_df = weather_df[weather_df['month'].isin(months)].copy()
                    
                    self.write_detailed_sheet(writer, seasonal_df, season)
                    self.write_aggregated_sheet(writer, seasonal_df, season)
                
                self.adjust_column_widths(writer.book)
            
            print("Real weather DataFrames written to Excel file with adjusted column widths.")
        finally:
            self.threadlock.release()


class PredictedWeatherWriter(BaseWeatherWriter):
    def run(self):
        self.threadlock.acquire()
        try:
            weather_df = self.output_queue.get()
            self.output_queue.put(weather_df.copy())
            print("Writing predicted weather data into Excel file with seasonal optimizations...")
            
            weather_df = self.prepare_data(weather_df)
            
            with pd.ExcelWriter('predicted_weather_data_seasonal.xlsx', engine='openpyxl') as writer:
                weather_df.to_excel(writer, sheet_name='Predicted Weather Data', index=False)
                
                for season, months in self.seasons.items():
                    seasonal_df = weather_df[weather_df['month'].isin(months)].copy()
                    
                    self.write_detailed_sheet(writer, seasonal_df, season)
                    self.write_aggregated_sheet(writer, seasonal_df, season)
                
                self.adjust_column_widths(writer.book)
            
            print("Predicted weather DataFrames written to Excel file with adjusted column widths.")
        finally:
            self.threadlock.release()



# class WeatherChartWriter(threading.Thread):
#     def __init__(self, real_queue, predicted_queue, threadlock):
#         print("[ChartWriter] Initializing weather chart writer...")
#         super().__init__()
#         self.real_queue = real_queue
#         self.predicted_queue = predicted_queue
#         self.threadlock = threadlock
        
#         self.variables = {
#             'Temperature': ('temperature', '°C'),  # Changed from 'temp' to 'temperature'
#             'Humidity': ('relative_humidity', '%'),
#             # 'Evaporation': ('evaporation', 'mm')  # Changed from 'evap' to 'evaporation'
#         }
        
#     def create_comparison_plot(self, x_data, y_data, x_label, title, year, is_predicted=False):
#         """Create a scatter plot for comparison"""
#         plt.figure(figsize=(8, 6))
#         color = 'orange' if is_predicted else 'blue'
#         label = f'{year} {"Predicted" if is_predicted else "Real"}'
#         plt.scatter(x_data, y_data, c=color, label=label)
        
#         plt.xlabel(x_label)
#         plt.ylabel('Precipitation (mm)')
#         plt.title(title)
#         plt.grid(True, linestyle='--', alpha=0.7)
#         plt.legend()
        
#         buf = BytesIO()
#         plt.savefig(buf, format='png', dpi=300, bbox_inches='tight')
#         plt.close()
#         buf.seek(0)
#         return buf

#     def create_merged_plot(self, real_x, real_y, pred_x, pred_y, x_label, title, year):
#         """Create a merged plot comparing real and predicted data"""
#         plt.figure(figsize=(8, 6))
        
#         plt.scatter(real_x, real_y, c='blue', label=f'{year} Real')
#         plt.scatter(pred_x, pred_y, c='orange', label=f'{year} Predicted')
        
#         plt.xlabel(x_label)
#         plt.ylabel('Precipitation (mm)')
#         plt.title(title)
#         plt.grid(True, linestyle='--', alpha=0.7)
#         plt.legend()

#         buf = BytesIO()
#         plt.savefig(buf, format='png', dpi=300, bbox_inches='tight')
#         plt.close()
#         buf.seek(0)
#         return buf

#     def add_chart_to_worksheet(self, worksheet: Worksheet, chart_buffer: BytesIO, 
#                             start_row: int, start_col: int, title: str):
#         """Add a chart to the worksheet at specified position"""
        
#         try:
#             img = Image(chart_buffer)
#             # Scale image
#             img.width = 500
#             img.height = 300
            
#             # Add title
#             worksheet.cell(row=start_row, column=start_col, value=title)
#             worksheet.cell(row=start_row, column=start_col).font = openpyxl.styles.Font(bold=True)
            
#             # Add image below title
#             worksheet.add_image(img, f'{chr(64 + start_col)}{start_row + 1}')
#             return start_row + 20  # Return next available row
#         except Exception as e:
#             raise

#     def run(self):

#         self.threadlock.acquire()
#         try:
#             real_df = self.real_queue.get()
#             predicted_df = self.predicted_queue.get()

#             with pd.ExcelWriter('weather_charts.xlsx', engine='openpyxl') as writer:
#                 real_sheet = writer.book.create_sheet(title='Charts (Real Data)')
#                 pred_sheet = writer.book.create_sheet(title='Charts (Predicted Data)')
#                 merged_sheet = writer.book.create_sheet(title='Charts (Merged)')

#                 years = sorted(real_df['date'].dt.year.unique())

#                 for year in years:
#                     real_year_data = real_df
#                     pred_year_data = predicted_df

#                     real_row = 1  # Track rows for the Real data sheet
#                     pred_row = 1  # Track rows for the Predicted data sheet
#                     merged_row = 1  # Track rows for the Merged sheet

#                     for var_name, (var_col, unit) in self.variables.items():

#                         # Create charts for real data
#                         title = f'Precipitation (Y) and {var_name} (X)'
#                         real_buf = self.create_comparison_plot(
#                             real_year_data[var_col],
#                             real_year_data['total_precipitation'],
#                             f'{var_name} ({unit})',
#                             title,
#                             year
#                         )
#                         real_row = self.add_chart_to_worksheet(
#                             real_sheet, real_buf, real_row, 1,
#                             f'{title} - {year}'
#                         )

#                         # Create charts for predicted data
#                         pred_buf = self.create_comparison_plot(
#                             pred_year_data[var_col],
#                             pred_year_data['total_precipitation'],
#                             f'{var_name} ({unit})',
#                             title,
#                             year,
#                             is_predicted=True
#                         )

#                         pred_row = self.add_chart_to_worksheet(
#                             pred_sheet, pred_buf, pred_row, 1,
#                             f'{title} - {year} (Predicted)'
#                         )

#                         merged_buf = self.create_merged_plot(
#                             real_year_data[var_col],
#                             real_year_data['total_precipitation'],
#                             pred_year_data[var_col],
#                             pred_year_data['total_precipitation'],
#                             f'{var_name} ({unit})',
#                             title,
#                             year
#                         )
#                         merged_row = self.add_chart_to_worksheet(
#                             merged_sheet, merged_buf, merged_row, 1,
#                             f'{title} - {year} (Merged)'
#                         )

            
#         except Exception as e:
#             print(f"[ChartWriter] Error in chart creation process: {str(e)}")
#             raise
#         finally:
#             print("[ChartWriter] Releasing thread lock")
#             self.threadlock.release()
#             print("[ChartWriter] Thread lock released")