import pandas as pd
import threading
import matplotlib.pyplot as plt
import ee
import numpy as np
from utils import calculate_relative_humidity, calculate_specific_humidity
from openpyxl.styles import PatternFill, Font
from openpyxl.utils.dataframe import dataframe_to_rows
import matplotlib.pyplot as plt
from io import BytesIO
from openpyxl.drawing.image import Image
from openpyxl.worksheet.worksheet import Worksheet
import pandas as pd
import openpyxl.styles


class WeatherAggregator(threading.Thread):
    def __init__(self, real_queue, predicted_queue, threadlock):
        super().__init__()
        self.real_queue = real_queue
        self.predicted_queue = predicted_queue
        self.threadlock = threadlock
        self.block_color = PatternFill(start_color="FFDDDDDD", end_color="FFDDDDDD", fill_type="solid")

    def create_yearly_aggregation(self, df, is_predicted=False):
        """Create yearly aggregation for the dataset"""
        df['year'] = pd.to_datetime(df['date']).dt.year if not is_predicted else pd.to_datetime(df['date']).dt.year
        
        variables = ['temperature', 'relative_humidity', 'total_precipitation'] if not is_predicted else ['total_precipitation']
        
        yearly_stats = {}
        for var in variables:
            yearly_stats[var] = df.groupby('year')[var].agg([
                ('mean', 'mean'),
                ('std', 'std'),
                ('min', 'min'),
                ('max', 'max')
            ]).round(2)
        
        return yearly_stats

    def create_comparison_sheet(self, writer, real_stats, pred_stats):
        """Create comparison sheet between real and predicted aggregated data"""
        workbook = writer.book
        worksheet = workbook.create_sheet(title='Real vs Predicted Comparison')
        
        # Write headers
        headers = ['Year', 'Variable', 'Type', 'Mean', 'Std', 'Min', 'Max']
        for col, header in enumerate(headers, start=1):
            worksheet.cell(row=1, column=col, value=header).font = Font(bold=True)
        
        current_row = 2
        variables = ['temperature', 'relative_humidity', 'total_precipitation']
        
        for year in real_stats['temperature'].index:
            for var in variables:
                # Write real data
                if var in real_stats:
                    real_data = real_stats[var].loc[year]
                    worksheet.cell(row=current_row, column=1, value=year)
                    worksheet.cell(row=current_row, column=2, value=var)
                    worksheet.cell(row=current_row, column=3, value='Real')
                    for col, value in enumerate(real_data, start=4):
                        cell = worksheet.cell(row=current_row, column=col, value=value)
                        cell.fill = self.block_color
                    current_row += 1
                
                # Write predicted data (only for precipitation)
                if var == 'total_precipitation' and year in pred_stats['total_precipitation'].index:
                    pred_data = pred_stats['total_precipitation'].loc[year]
                    worksheet.cell(row=current_row, column=1, value=year)
                    worksheet.cell(row=current_row, column=2, value=var)
                    worksheet.cell(row=current_row, column=3, value='Predicted')
                    for col, value in enumerate(pred_data, start=4):
                        cell = worksheet.cell(row=current_row, column=col, value=value)
                        cell.fill = PatternFill(start_color="FFFFCC99", end_color="FFFFCC99", fill_type="solid")
                    current_row += 1
            
            current_row += 1  # Add space between years
        
        self.adjust_column_widths(worksheet)

    def adjust_column_widths(self, worksheet):
        """Adjust column widths for better readability"""
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = max_length + 2
            worksheet.column_dimensions[column_letter].width = adjusted_width

    def run(self):
        self.threadlock.acquire()
        try:
            print("Starting weather aggregation process...")
            
            # Get data from queues
            real_df = self.real_queue.get()
            predicted_df = self.predicted_queue.get()
            
            # Put the data back in the queues for other threads
            self.real_queue.put(real_df)
            self.predicted_queue.put(predicted_df)
            
            # Create yearly aggregations
            real_stats = self.create_yearly_aggregation(real_df, is_predicted=False)
            pred_stats = self.create_yearly_aggregation(predicted_df, is_predicted=True)
            
            # Create Excel file with comparisons
            with pd.ExcelWriter('weather_aggregated_comparison.xlsx', engine='openpyxl') as writer:
                # Write individual sheets for real data
                for var, stats in real_stats.items():
                    sheet_name = f'Real_{var}_yearly'
                    stats.to_excel(writer, sheet_name=sheet_name)
                
                # Write precipitation predictions
                pred_stats['total_precipitation'].to_excel(writer, sheet_name='Predicted_precipitation_yearly')
                
                # Create comparison sheet
                self.create_comparison_sheet(writer, real_stats, pred_stats)
            
            print("Weather aggregation complete and saved to weather_aggregated_comparison.xlsx")
            
        finally:
            self.threadlock.release()